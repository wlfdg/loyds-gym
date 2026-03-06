from flask import Flask, request, jsonify, make_response
from flask_cors import CORS
from datetime import datetime, timedelta
import psycopg2
import psycopg2.extras
import os
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

DATABASE_URL = os.environ.get("DATABASE_URL")

def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    return conn

def dict_cur(conn):
    return conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS admins (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS members (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT,
            phone TEXT,
            plan TEXT NOT NULL,
            months INTEGER NOT NULL,
            price REAL NOT NULL,
            discount REAL DEFAULT 0,
            start_date TEXT NOT NULL,
            expiration_date TEXT NOT NULL,
            status TEXT DEFAULT 'active'
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS walkins (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            amount REAL NOT NULL,
            note TEXT,
            date TEXT NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id SERIAL PRIMARY KEY,
            member_id INTEGER NOT NULL,
            member_name TEXT NOT NULL,
            time_in TEXT NOT NULL,
            time_out TEXT,
            date TEXT NOT NULL
        )
    """)
    cur.execute("SELECT COUNT(*) FROM admins")
    if cur.fetchone()[0] == 0:
        cur.execute("INSERT INTO admins (username, password) VALUES (%s, %s)", ("admin", "admin123"))
    conn.commit()
    cur.close()
    conn.close()

init_db()

# ── Auth ──────────────────────────────────────────────────────────────────────

@app.route("/login", methods=["POST"])
def login():
    data = request.json
    conn = get_db()
    cur = dict_cur(conn)
    cur.execute("SELECT * FROM admins WHERE username=%s AND password=%s", (data["username"], data["password"]))
    admin = cur.fetchone()
    cur.close()
    conn.close()
    if admin:
        return jsonify({"success": True, "username": admin["username"]})
    return jsonify({"success": False, "message": "Invalid credentials"}), 401

@app.route("/register", methods=["POST"])
def register():
    data = request.json
    if not data or not data.get("username") or not data.get("password"):
        return jsonify({"message": "Username and password required"}), 400
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("INSERT INTO admins (username, password) VALUES (%s, %s)",
                    (data["username"].strip(), data["password"].strip()))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"message": "Admin created", "success": True})
    except psycopg2.IntegrityError:
        return jsonify({"message": "Username already exists", "success": False}), 409

# ── Members ───────────────────────────────────────────────────────────────────

@app.route("/members", methods=["GET"])
def get_members():
    conn = get_db()
    cur = dict_cur(conn)
    cur.execute("SELECT * FROM members ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/members/<int:id>", methods=["GET"])
def get_member(id):
    conn = get_db()
    cur = dict_cur(conn)
    cur.execute("SELECT * FROM members WHERE id=%s", (id,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if row:
        return jsonify(dict(row))
    return jsonify({"message": "Not found"}), 404

@app.route("/members", methods=["POST"])
def add_member():
    data = request.json
    months = int(data["months"])
    start = datetime.now()
    expiration = start + timedelta(days=30 * months)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO members (name, email, phone, plan, months, price, discount, start_date, expiration_date, status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'active')
    """, (
        data["name"], data.get("email", ""), data.get("phone", ""),
        data["plan"], months, float(data["price"]), float(data.get("discount", 0)),
        start.strftime("%Y-%m-%d"), expiration.strftime("%Y-%m-%d")
    ))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"message": "Member added"}), 201

@app.route("/members/<int:id>", methods=["PUT"])
def update_member(id):
    data = request.json
    months = int(data["months"])
    start = data.get("start_date", datetime.now().strftime("%Y-%m-%d"))
    expiration = (datetime.strptime(start, "%Y-%m-%d") + timedelta(days=30 * months)).strftime("%Y-%m-%d")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE members SET name=%s, email=%s, phone=%s, plan=%s, months=%s, price=%s, discount=%s,
        start_date=%s, expiration_date=%s WHERE id=%s
    """, (
        data["name"], data.get("email", ""), data.get("phone", ""),
        data["plan"], months, float(data["price"]), float(data.get("discount", 0)),
        start, expiration, id
    ))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"message": "Member updated"})

@app.route("/members/<int:id>", methods=["DELETE"])
def delete_member(id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM members WHERE id=%s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"message": "Deleted"})

# ── Attendance ────────────────────────────────────────────────────────────────

@app.route("/attendance/timein", methods=["POST"])
def time_in():
    data = request.json
    member_id = data["member_id"]
    now = datetime.now()
    today = now.strftime("%Y-%m-%d")
    time_now = now.strftime("%I:%M %p")
    conn = get_db()
    cur = dict_cur(conn)
    cur.execute("SELECT * FROM attendance WHERE member_id=%s AND date=%s AND time_out IS NULL", (member_id, today))
    existing = cur.fetchone()
    if existing:
        cur.close()
        conn.close()
        return jsonify({"message": "Member is already timed in today."}), 409
    cur.execute("SELECT name FROM members WHERE id=%s", (member_id,))
    member = cur.fetchone()
    if not member:
        cur.close()
        conn.close()
        return jsonify({"message": "Member not found"}), 404
    cur2 = conn.cursor()
    cur2.execute("INSERT INTO attendance (member_id, member_name, time_in, date) VALUES (%s, %s, %s, %s)",
                 (member_id, member["name"], time_now, today))
    conn.commit()
    cur.close()
    cur2.close()
    conn.close()
    return jsonify({"message": f"{member['name']} timed in at {time_now}"}), 201

@app.route("/attendance/timeout", methods=["POST"])
def time_out():
    data = request.json
    member_id = data["member_id"]
    now = datetime.now()
    today = now.strftime("%Y-%m-%d")
    time_now = now.strftime("%I:%M %p")
    conn = get_db()
    cur = dict_cur(conn)
    cur.execute("SELECT * FROM attendance WHERE member_id=%s AND date=%s AND time_out IS NULL", (member_id, today))
    record = cur.fetchone()
    if not record:
        cur.close()
        conn.close()
        return jsonify({"message": "No active time-in found for this member today."}), 404
    cur2 = conn.cursor()
    cur2.execute("UPDATE attendance SET time_out=%s WHERE id=%s", (time_now, record["id"]))
    conn.commit()
    cur.close()
    cur2.close()
    conn.close()
    return jsonify({"message": f"Timed out at {time_now}"}), 200

@app.route("/attendance", methods=["GET"])
def get_attendance():
    date = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    conn = get_db()
    cur = dict_cur(conn)
    cur.execute("SELECT * FROM attendance WHERE date=%s ORDER BY id DESC", (date,))
    rows = cur.fetchall()
    cur.execute("SELECT COUNT(*) as c FROM attendance WHERE date=%s AND time_out IS NULL", (date,))
    still_in = cur.fetchone()["c"]
    cur.close()
    conn.close()
    return jsonify({"records": [dict(r) for r in rows], "total": len(rows), "still_in": still_in, "date": date})

@app.route("/attendance/history/<int:member_id>", methods=["GET"])
def member_attendance(member_id):
    conn = get_db()
    cur = dict_cur(conn)
    cur.execute("SELECT * FROM attendance WHERE member_id=%s ORDER BY id DESC LIMIT 30", (member_id,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ── Walk-ins ──────────────────────────────────────────────────────────────────

@app.route("/walkins", methods=["GET"])
def get_walkins():
    date = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    conn = get_db()
    cur = dict_cur(conn)
    cur.execute("SELECT * FROM walkins WHERE date=%s ORDER BY id DESC", (date,))
    rows = cur.fetchall()
    cur.execute("SELECT SUM(amount) as t FROM walkins WHERE date=%s", (date,))
    result = cur.fetchone()
    total = result["t"] or 0 if result else 0
    cur.close()
    conn.close()
    return jsonify({"walkins": [dict(r) for r in rows], "total": round(float(total), 2), "date": date})

@app.route("/walkins", methods=["POST"])
def add_walkin():
    data = request.json
    today = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO walkins (name, amount, note, date) VALUES (%s, %s, %s, %s)",
                (data["name"], float(data["amount"]), data.get("note", ""), today))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"message": "Walk-in recorded"}), 201

@app.route("/walkins/<int:id>", methods=["DELETE"])
def delete_walkin(id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM walkins WHERE id=%s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"message": "Deleted"})

# ── Stats ─────────────────────────────────────────────────────────────────────

@app.route("/stats", methods=["GET"])
def stats():
    today = datetime.now().strftime("%Y-%m-%d")
    now_ym = datetime.now().strftime("%Y-%m")
    conn = get_db()
    cur = dict_cur(conn)

    cur.execute("SELECT COUNT(*) as c FROM members")
    total = cur.fetchone()["c"]

    cur.execute("SELECT COUNT(*) as c FROM members WHERE expiration_date >= %s", (today,))
    active = cur.fetchone()["c"]

    cur.execute("SELECT SUM(price - (price * discount / 100)) as r FROM members")
    rev = cur.fetchone()["r"] or 0

    cur.execute("SELECT COUNT(*) as c FROM members WHERE TO_CHAR(TO_DATE(start_date,'YYYY-MM-DD'),'YYYY-MM') = %s", (now_ym,))
    new_this_month = cur.fetchone()["c"]

    cur.execute("SELECT SUM(amount) as t FROM walkins WHERE date=%s", (today,))
    walkin_today = cur.fetchone()["t"] or 0

    cur.execute("SELECT COUNT(*) as c FROM walkins WHERE date=%s", (today,))
    walkin_count_today = cur.fetchone()["c"]

    cur.execute("SELECT COUNT(*) as c FROM attendance WHERE date=%s AND time_out IS NULL", (today,))
    members_in_gym = cur.fetchone()["c"]

    cur.execute("SELECT COUNT(*) as c FROM attendance WHERE date=%s", (today,))
    visits_today = cur.fetchone()["c"]

    cur.close()
    conn.close()
    return jsonify({
        "total_members": total,
        "active_members": active,
        "revenue": round(float(rev), 2),
        "new_this_month": new_this_month,
        "walkin_revenue_today": round(float(walkin_today), 2),
        "walkin_count_today": walkin_count_today,
        "members_in_gym": members_in_gym,
        "visits_today": visits_today
    })

# ── Expiring ──────────────────────────────────────────────────────────────────

@app.route("/expiring", methods=["GET"])
def expiring():
    days = int(request.args.get("days", 7))
    future = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d")
    today = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    cur = dict_cur(conn)
    cur.execute("SELECT * FROM members WHERE expiration_date <= %s AND expiration_date >= %s ORDER BY expiration_date ASC", (future, today))
    rows = cur.fetchall()
    cur.execute("SELECT * FROM members WHERE expiration_date < %s ORDER BY expiration_date DESC", (today,))
    expired = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify({"expiring_soon": [dict(r) for r in rows], "expired": [dict(r) for r in expired]})

# ── Export CSV ────────────────────────────────────────────────────────────────

@app.route("/export/csv", methods=["GET"])
def export_csv():
    conn = get_db()
    cur = dict_cur(conn)
    cur.execute("SELECT * FROM members ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Name", "Email", "Phone", "Plan", "Months", "Price", "Discount%", "Start Date", "Expiration Date"])
    for r in rows:
        writer.writerow([r["id"], r["name"], r["email"], r["phone"], r["plan"],
                         r["months"], r["price"], r["discount"], r["start_date"], r["expiration_date"]])
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=members.csv"
    response.headers["Content-Type"] = "text/csv"
    return response

# ── Monthly Report Excel ──────────────────────────────────────────────────────

@app.route("/report/excel", methods=["GET"])
def report_excel():
    year = request.args.get("year", datetime.now().strftime("%Y"))
    month = request.args.get("month", datetime.now().strftime("%m"))
    month_str = f"{year}-{month.zfill(2)}"
    month_name = datetime.strptime(month_str, "%Y-%m").strftime("%B %Y")

    conn = get_db()
    cur = dict_cur(conn)

    cur.execute("SELECT * FROM members WHERE TO_CHAR(TO_DATE(start_date,'YYYY-MM-DD'),'YYYY-MM') = %s", (month_str,))
    new_members = cur.fetchall()

    cur.execute("SELECT * FROM members")
    all_members = cur.fetchall()

    today = datetime.now().strftime("%Y-%m-%d")
    cur.execute("SELECT * FROM members WHERE expiration_date >= %s", (today,))
    active_members = cur.fetchall()

    cur.execute("SELECT * FROM walkins WHERE TO_CHAR(TO_DATE(date,'YYYY-MM-DD'),'YYYY-MM') = %s ORDER BY date ASC", (month_str,))
    walkins = cur.fetchall()

    cur.execute("""
        SELECT a.*, m.plan FROM attendance a
        LEFT JOIN members m ON a.member_id = m.id
        WHERE TO_CHAR(TO_DATE(a.date,'YYYY-MM-DD'),'YYYY-MM') = %s
        ORDER BY a.date ASC, a.time_in ASC
    """, (month_str,))
    attendance = cur.fetchall()

    cur.close()
    conn.close()

    member_revenue = sum((r["price"] - r["price"] * r["discount"] / 100) for r in new_members)
    walkin_revenue = sum(float(r["amount"]) for r in walkins)
    total_revenue = member_revenue + walkin_revenue

    DARK = "1a1a1a"
    YELLOW = "E8FF00"
    HEADER_BG = "2a2a2a"
    HEADER_FG = "E8FF00"
    WHITE = "F0F0F0"
    GREEN = "00E676"
    MUTED = "888888"
    ROW_ALT = "1f1f1f"

    thin = Side(style="thin", color="2a2a2a")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    def style_header_row(ws, row, cols, bg=HEADER_BG, fg=HEADER_FG):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, color=fg, name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def style_data_row(ws, row, cols, alt=False):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(color=WHITE, name="Arial", size=9)
            cell.fill = PatternFill("solid", start_color=ROW_ALT if alt else DARK)
            cell.alignment = Alignment(vertical="center")
            cell.border = border

    def add_section_title(ws, row, col, text):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color=YELLOW, name="Arial", size=12)
        cell.fill = PatternFill("solid", start_color=DARK)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 24

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # SHEET 1 — Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False
    for row in ws1.iter_rows(min_row=1, max_row=60, min_col=1, max_col=10):
        for cell in row:
            cell.fill = PatternFill("solid", start_color=DARK)

    ws1.merge_cells("A1:H1")
    t = ws1["A1"]
    t.value = "LOYD'S FITNESS GYM"
    t.font = Font(bold=True, color=YELLOW, name="Arial", size=20)
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.fill = PatternFill("solid", start_color=DARK)
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:H2")
    s = ws1["A2"]
    s.value = f"Monthly Report — {month_name}"
    s.font = Font(color=WHITE, name="Arial", size=13)
    s.alignment = Alignment(horizontal="center", vertical="center")
    s.fill = PatternFill("solid", start_color=DARK)
    ws1.row_dimensions[2].height = 24

    ws1.merge_cells("A3:H3")
    g = ws1["A3"]
    g.value = f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
    g.font = Font(color=MUTED, name="Arial", size=9)
    g.alignment = Alignment(horizontal="center")
    g.fill = PatternFill("solid", start_color=DARK)
    ws1.row_dimensions[3].height = 18
    ws1.row_dimensions[4].height = 12

    kpis = [
        ("TOTAL MEMBERS", len(all_members), WHITE),
        ("ACTIVE MEMBERS", len(active_members), GREEN),
        ("NEW THIS MONTH", len(new_members), YELLOW),
        ("WALK-INS", len(walkins), "00B0FF"),
        ("ATTENDANCE LOGS", len(attendance), "FF9100"),
        ("MEMBER REVENUE", f"P{member_revenue:,.2f}", GREEN),
        ("WALKIN REVENUE", f"P{walkin_revenue:,.2f}", YELLOW),
        ("TOTAL REVENUE", f"P{total_revenue:,.2f}", "FF4D00"),
    ]

    ws1.merge_cells("A5:H5")
    kpi_title = ws1["A5"]
    kpi_title.value = "KEY PERFORMANCE INDICATORS"
    kpi_title.font = Font(bold=True, color=MUTED, name="Arial", size=9)
    kpi_title.alignment = Alignment(horizontal="left")
    kpi_title.fill = PatternFill("solid", start_color=DARK)
    ws1.row_dimensions[5].height = 18

    col = 1
    for label, value, color in kpis:
        lc = ws1.cell(row=6, column=col, value=label)
        lc.font = Font(color=MUTED, name="Arial", size=8, bold=True)
        lc.fill = PatternFill("solid", start_color=HEADER_BG)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = border
        ws1.row_dimensions[6].height = 20
        vc = ws1.cell(row=7, column=col, value=value)
        vc.font = Font(bold=True, color=color, name="Arial", size=14)
        vc.fill = PatternFill("solid", start_color=HEADER_BG)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = border
        ws1.row_dimensions[7].height = 32
        col += 1

    ws1.row_dimensions[8].height = 16
    add_section_title(ws1, 9, 1, f"NEW MEMBERS THIS MONTH ({len(new_members)})")
    headers = ["#", "Name", "Email", "Phone", "Plan", "Months", "Price", "Discount %", "Start Date", "Expiration"]
    for i, h in enumerate(headers, 1):
        ws1.cell(row=10, column=i, value=h)
    style_header_row(ws1, 10, len(headers))
    ws1.row_dimensions[10].height = 20

    for idx, m in enumerate(new_members):
        r = 11 + idx
        net = m["price"] - (m["price"] * m["discount"] / 100)
        row_data = [idx+1, m["name"], m["email"], m["phone"], m["plan"], m["months"], net, f'{m["discount"]}%', m["start_date"], m["expiration_date"]]
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)
        style_data_row(ws1, r, len(headers), alt=idx % 2 == 1)
        ws1.row_dimensions[r].height = 18

    if not new_members:
        ws1.merge_cells("A11:J11")
        empty = ws1["A11"]
        empty.value = "No new members this month."
        empty.font = Font(color=MUTED, name="Arial", size=9, italic=True)
        empty.fill = PatternFill("solid", start_color=DARK)
        empty.alignment = Alignment(horizontal="center")

    set_col_widths(ws1, [4, 20, 22, 14, 12, 8, 12, 10, 12, 12])

    # SHEET 2 — All Members
    ws2 = wb.create_sheet("All Members")
    ws2.sheet_view.showGridLines = False
    for row in ws2.iter_rows(min_row=1, max_row=len(all_members)+10, min_col=1, max_col=11):
        for cell in row:
            cell.fill = PatternFill("solid", start_color=DARK)

    ws2.merge_cells("A1:K1")
    t2 = ws2["A1"]
    t2.value = f"ALL MEMBERS — {month_name}"
    t2.font = Font(bold=True, color=YELLOW, name="Arial", size=14)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    t2.fill = PatternFill("solid", start_color=DARK)
    ws2.row_dimensions[1].height = 28

    headers2 = ["#", "Name", "Email", "Phone", "Plan", "Months", "Net Price", "Discount %", "Start Date", "Expiration", "Status"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=i, value=h)
    style_header_row(ws2, 2, len(headers2))

    today_str = datetime.now().strftime("%Y-%m-%d")
    for idx, m in enumerate(all_members):
        r = 3 + idx
        net = m["price"] - (m["price"] * m["discount"] / 100)
        exp = m["expiration_date"]
        if exp < today_str:
            status = "Expired"
        elif (datetime.strptime(exp, "%Y-%m-%d") - datetime.now()).days <= 7:
            status = "Expiring Soon"
        else:
            status = "Active"
        row_data = [idx+1, m["name"], m["email"], m["phone"], m["plan"], m["months"], net, f'{m["discount"]}%', m["start_date"], exp, status]
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)
        style_data_row(ws2, r, len(row_data), alt=idx % 2 == 1)
        sc = ws2.cell(row=r, column=11)
        if status == "Active":
            sc.font = Font(color=GREEN, name="Arial", size=9, bold=True)
        elif status == "Expiring Soon":
            sc.font = Font(color="FFB300", name="Arial", size=9, bold=True)
        else:
            sc.font = Font(color="FF1744", name="Arial", size=9, bold=True)
        ws2.row_dimensions[r].height = 18

    set_col_widths(ws2, [4, 20, 22, 14, 12, 8, 14, 10, 12, 12, 14])

    # SHEET 3 — Walk-ins
    ws3 = wb.create_sheet("Walk-ins")
    ws3.sheet_view.showGridLines = False
    for row in ws3.iter_rows(min_row=1, max_row=len(walkins)+10, min_col=1, max_col=5):
        for cell in row:
            cell.fill = PatternFill("solid", start_color=DARK)

    ws3.merge_cells("A1:E1")
    t3 = ws3["A1"]
    t3.value = f"WALK-IN REVENUE — {month_name}"
    t3.font = Font(bold=True, color=YELLOW, name="Arial", size=14)
    t3.alignment = Alignment(horizontal="center", vertical="center")
    t3.fill = PatternFill("solid", start_color=DARK)
    ws3.row_dimensions[1].height = 28

    headers3 = ["#", "Name", "Amount", "Note", "Date"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=2, column=i, value=h)
    style_header_row(ws3, 2, len(headers3))

    for idx, w in enumerate(walkins):
        r = 3 + idx
        row_data = [idx+1, w["name"], float(w["amount"]), w["note"] or "—", w["date"]]
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)
        style_data_row(ws3, r, len(row_data), alt=idx % 2 == 1)
        ws3.cell(row=r, column=3).font = Font(color=YELLOW, name="Arial", size=9, bold=True)
        ws3.row_dimensions[r].height = 18

    total_row = 3 + len(walkins) + 1
    tc = ws3.cell(row=total_row, column=2, value="TOTAL")
    tc.font = Font(bold=True, color=WHITE, name="Arial", size=10)
    tc.fill = PatternFill("solid", start_color=HEADER_BG)
    tv = ws3.cell(row=total_row, column=3, value=walkin_revenue)
    tv.font = Font(bold=True, color=YELLOW, name="Arial", size=12)
    tv.fill = PatternFill("solid", start_color=HEADER_BG)
    ws3.row_dimensions[total_row].height = 24
    set_col_widths(ws3, [4, 22, 14, 24, 14])

    # SHEET 4 — Attendance
    ws4 = wb.create_sheet("Attendance")
    ws4.sheet_view.showGridLines = False
    for row in ws4.iter_rows(min_row=1, max_row=len(attendance)+10, min_col=1, max_col=7):
        for cell in row:
            cell.fill = PatternFill("solid", start_color=DARK)

    ws4.merge_cells("A1:G1")
    t4 = ws4["A1"]
    t4.value = f"ATTENDANCE LOG — {month_name}"
    t4.font = Font(bold=True, color=YELLOW, name="Arial", size=14)
    t4.alignment = Alignment(horizontal="center", vertical="center")
    t4.fill = PatternFill("solid", start_color=DARK)
    ws4.row_dimensions[1].height = 28

    headers4 = ["#", "Member", "Plan", "Date", "Time In", "Time Out", "Status"]
    for i, h in enumerate(headers4, 1):
        ws4.cell(row=2, column=i, value=h)
    style_header_row(ws4, 2, len(headers4))

    for idx, a in enumerate(attendance):
        r = 3 + idx
        status = "Inside" if not a["time_out"] else "Left"
        row_data = [idx+1, a["member_name"], a["plan"] or "—", a["date"], a["time_in"], a["time_out"] or "—", status]
        for c, val in enumerate(row_data, 1):
            ws4.cell(row=r, column=c, value=val)
        style_data_row(ws4, r, len(row_data), alt=idx % 2 == 1)
        sc = ws4.cell(row=r, column=7)
        sc.font = Font(color=GREEN if status == "Inside" else "FF1744", name="Arial", size=9, bold=True)
        ws4.row_dimensions[r].height = 18

    set_col_widths(ws4, [4, 22, 14, 12, 12, 12, 10])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"LoydsGym_Report_{month_name.replace(' ', '_')}.xlsx"
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
